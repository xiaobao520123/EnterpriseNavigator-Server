import cfg

import os
import re
import requests
from werkzeug.utils import secure_filename

import timeit
import time
import openpyxl

import urllib.request
from tqdm import tqdm

from PIL import Image, ImageDraw

# 检查是不是允许上传的文件类型
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in cfg.allowed_extenstions

# 转存上传的图片文件
def save_uploaded_file(raw_files):
    for file in raw_files:
        if file:
            filename = file.filename
            if not allowed_file(filename):
                continue
            filename = secure_filename(filename)
            img_path = os.path.join(cfg.image_preview_folder, filename)
            file.save(img_path)   

# 转发批量文件上传请求到计算识别服务器
def batch_upload(headers, file_data):
    start = timeit.default_timer()

    response = requests.post(url=cfg.api_qcc_batch_detection, 
        headers=headers, 
        data=file_data)
    
    end = timeit.default_timer()
    delay = end-start
    print('Eclapsed： %s 秒' % delay)
    return response, delay

# 将批量识别的结果保存到表格文件
def write_batch_detection_result(collections):
    # 以当前时间作为表格文件名
    now = time.strftime("%Y-%m-%d-%H-%M-%S",time.localtime(time.time())) 
    # 获得表格文件输出路径
    output_filepath = os.path.join(cfg.table_output_folder, now + '.xlsx')

    # 创建Excel工作本
    book = openpyxl.Workbook()
    # 创建组
    sheet = book.create_sheet('识别结果', 0)

    sheet.cell(1, 1, '图片')
    sheet.cell(1, 2, '商铺名称')    
    # 将结果逐条写入表格中
    for i in range(len(collections)):
        sheet.cell(i + 1 + 1, 1, collections[i]['filename'])
        sheet.cell(i + 1 + 1, 2, collections[i]['enterprise'])
    
    # 保存表格文件到路径
    book.save(output_filepath)
    return output_filepath

# 在图片上绘制结果框
def draw_polygon_on_image(img_path, polygon):
    image = Image.open(img_path).convert('RGB') # 转换色彩空间
    draw = ImageDraw.Draw(image)
    draw.line([tuple(polygon[0]),
                tuple(polygon[1]),
                tuple(polygon[2]),
                tuple(polygon[3]),
                tuple(polygon[0])], width=cfg.answer_line_width, fill=cfg.answer_line_color)
    image.save(img_path)

# 排序用比较函数，先比较文件名中的数字大小，再比较文件名长度
# <<需要加强！>>
def sort_by_filename(elem):
    digit_str = re.sub(r'\D', '', elem['filename'])
    if len(digit_str) == 0:
        return len(elem['filename'])
    return int(digit_str)

def process_detection_result(json_data):
    collections = []
    items = json_data.items()
    for key, value in items:
        if value is None:
            value = ''

        detection_result = value
        success = detection_result['success']
        enterprise = detection_result['enterprise']
        polygon = detection_result['polygon']

        filename = key

        # 处理文件名，去除根目录名
        slash_index = filename.find('/')
        if slash_index != -1:
            filename = filename[slash_index+1:]
        
        # 识别失败
        if success == 0:
            enterprise = '<识别失败>'

        img_filename = secure_filename(key)
        img_path = os.path.join(cfg.image_preview_folder, img_filename)

        if polygon and len(polygon) > 0:
            draw_polygon_on_image(img_path, polygon)

        collection = {
            "filename": filename,
            "enterprise": enterprise,
            "polygon": polygon,
            "img_preview": img_path,
        }
        collections.append(collection)
    if len(collections) > 1:
        collections.sort(key=sort_by_filename)
    return collections

class DownloadProgressBar(tqdm):
    def update_to(self, b=1, bsize=1, tsize=None):
        if tsize is not None:
            self.total = tsize
        self.update(b * bsize - self.n)


def download_url(url, output_path):
    with DownloadProgressBar(unit='B', unit_scale=True,
                             miniters=1, desc=url.split('/')[-1]) as t:
        urllib.request.urlretrieve(url, filename=output_path, reporthook=t.update_to)